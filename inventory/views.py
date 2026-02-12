from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, FileResponse, Http404
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Maintenance, Handover, Equipment, Peripheral, Area, CostCenter, MaintenanceSchedule
from .utils import generate_maintenance_pdf, generate_handover_pdf, export_to_excel
from django.apps import apps
from .forms import MaintenanceForm, EquipmentForm, AreaForm, CostCenterForm, CustomUserCreationForm, PeripheralForm, HandoverForm, ClientForm
from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from urllib.parse import urlencode
from django.urls import reverse
import qrcode
import openpyxl
from .forms import ExcelImportForm

@login_required
def maintenance_acta_view(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)
    
    if maintenance.acta_pdf:
        return FileResponse(maintenance.acta_pdf, as_attachment=False, filename=f"acta_mantenimiento_{pk}.pdf")
    
    # Fallback if PDF not generated yet
    pdf_content = generate_maintenance_pdf(maintenance)
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="acta_mantenimiento_{pk}.pdf"'
    return response

@login_required
def handover_acta_view(request, pk):
    handover = get_object_or_404(Handover, pk=pk)
    
    if handover.acta_pdf:
        return FileResponse(handover.acta_pdf, as_attachment=False, filename=f"acta_entrega_{pk}.pdf")

    pdf_content = generate_handover_pdf(handover)
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="acta_entrega_{pk}.pdf"'
    return response

@login_required
def export_data_view(request, model_name):
    allowed_models = ['equipment', 'peripheral', 'maintenance', 'handover', 'client', 'area', 'costcenter']
    if model_name.lower() not in allowed_models:
        raise Http404("Invalid model for export")

    try:
        model = apps.get_model('inventory', model_name)
    except LookupError:
        raise Http404("Model not found")
    
    queryset = model.objects.all()

    # Apply filters based on model name
    if model_name == 'equipment':
        query = request.GET.get('q', '')
        area_id = request.GET.get('area', '')
        status = request.GET.get('status', '')
        eq_type = request.GET.get('type', '')
        ownership = request.GET.get('ownership', '')

        if query:
            queryset = queryset.filter(
                Q(serial_number__icontains=query) | 
                Q(brand__icontains=query) | 
                Q(model__icontains=query) |
                Q(ip_address__icontains=query)
            )
        if area_id:
            queryset = queryset.filter(area_id=area_id)
        if status:
            queryset = queryset.filter(status=status)
        if eq_type:
            queryset = queryset.filter(type=eq_type)
        if ownership:
            queryset = queryset.filter(ownership_type=ownership)
            
    elif model_name == 'maintenance':
        date_start = request.GET.get('date_start', '')
        date_end = request.GET.get('date_end', '')
        m_type = request.GET.get('type', '')
        
        if date_start:
            queryset = queryset.filter(date__gte=date_start)
        if date_end:
            queryset = queryset.filter(date__lte=date_end)
        if m_type:
            queryset = queryset.filter(maintenance_type=m_type)
            
    elif model_name == 'handover':
        date_start = request.GET.get('date_start', '')
        date_end = request.GET.get('date_end', '')
        area_id = request.GET.get('area', '')
        
        if date_start:
            queryset = queryset.filter(date__date__gte=date_start)
        if date_end:
            queryset = queryset.filter(date__date__lte=date_end)
        if area_id:
            queryset = queryset.filter(Q(source_area_id=area_id) | Q(destination_area_id=area_id))
            
    elif model_name == 'peripheral':
        query = request.GET.get('q', '')
        if query:
            queryset = queryset.filter(
                Q(serial_number__icontains=query) | 
                Q(brand__icontains=query) | 
                Q(model__icontains=query) |
                Q(type__icontains=query)
            )

    # Create a dummy ModelAdmin-like object or just pass a simple object with model._meta
    class DummyAdmin:
        pass
    dummy_admin = DummyAdmin()
    dummy_admin.model = model
    
    return export_to_excel(queryset, dummy_admin, request)

@login_required
def dashboard_view(request):
    total_equipment = Equipment.objects.count()
    active_equipment = Equipment.objects.filter(status='ACTIVE').count()
    maintenance_count = Maintenance.objects.count()
    
    dashboard_view_handovers = Handover.objects.count()
    
    recent_maintenance = Maintenance.objects.order_by('-date')[:5]
    recent_handovers = Handover.objects.order_by('-date')[:5]
    
    # Chart Data Preparation
    from .choices import EQUIPMENT_STATUS_CHOICES, EQUIPMENT_TYPE_CHOICES
    status_dict = dict(EQUIPMENT_STATUS_CHOICES)
    type_dict = dict(EQUIPMENT_TYPE_CHOICES)

    raw_status_data = Equipment.objects.values('status').annotate(count=Count('status'))
    status_data = [{'status': status_dict.get(item['status'], item['status']), 'count': item['count']} for item in raw_status_data]

    raw_type_data = Equipment.objects.values('type').annotate(count=Count('type'))
    type_data = [{'type': type_dict.get(item['type'], item['type']), 'count': item['count']} for item in raw_type_data]

    area_data = list(Equipment.objects.exclude(area__isnull=True).values('area__name').annotate(count=Count('id')).order_by('-count')[:5])

    context = {
        'total_equipment': total_equipment,
        'active_equipment': active_equipment,
        'maintenance_count': maintenance_count,
        'handover_count': dashboard_view_handovers,
        'recent_maintenance': recent_maintenance,
        'recent_handovers': recent_handovers,
        'status_data': status_data,
        'type_data': type_data,
        'area_data': area_data,
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required
def generate_qr_view(request, pk):
    # Just verify existence
    equipment = get_object_or_404(Equipment, pk=pk)
    
    # URL to the detail page (absolute)
    url = request.build_absolute_uri(reverse('inventory:equipment_detail', args=[pk]))
    
    # Construct data as URL parameters so it opens automatically
    # but still contains readable info in the URL string
    params = {
        'serial': equipment.serial_number,
        'model': f"{equipment.brand} {equipment.model}",
        'type': equipment.get_type_display(),
        'area': equipment.area.name if equipment.area else 'N/A',
        'status': equipment.get_status_display()
    }
    
    full_url = f"{url}?{urlencode(params)}"

    # Create QR with higher error correction
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(full_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response

@login_required
def import_equipment_view(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(request.FILES['excel_file'])
                ws = wb.active
                
                # Assume headers are row 1, data starts row 2
                imported_count = 0
                errors = []
                
                for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Format: Serial (0), Type (1), Brand (2), Model (3), Area (4), Status (5)
                    serial = str(row[0]).strip() if row[0] else None
                    if not serial: continue # Skip empty rows
                    
                    eq_type = str(row[1]).upper() if row[1] else 'PC'
                    brand = str(row[2]).strip() if row[2] else 'Genérico'
                    model = str(row[3]).strip() if row[3] else 'Genérico'
                    area_name = str(row[4]).strip() if row[4] else None
                    status = str(row[5]).upper() if row[5] else 'ACTIVE'
                    
                    # Resolve Area
                    area = None
                    if area_name:
                        area, _ = Area.objects.get_or_create(name=area_name)
                    
                    # Create or Update
                    obj, created = Equipment.objects.update_or_create(
                        serial_number=serial,
                        defaults={
                            'type': eq_type,
                            'brand': brand,
                            'model': model,
                            'area': area,
                            'status': status
                        }
                    )
                    imported_count += 1
                
                return render(request, 'inventory/import_success.html', {'count': imported_count})
            
            except Exception as e:
                form.add_error(None, f"Error procesando el archivo: {str(e)}")
    else:
        form = ExcelImportForm()
        
    return render(request, 'inventory/import_form.html', {'form': form, 'title': 'Importar Equipos'})

@login_required
def equipment_list_view(request):
    query = request.GET.get('q', '')
    area_id = request.GET.get('area', '')
    status = request.GET.get('status', '')
    eq_type = request.GET.get('type', '')
    ownership = request.GET.get('ownership', '')
    
    equipments_list = Equipment.objects.all().order_by('-created_at')
    
    if query:
        equipments_list = equipments_list.filter(
            Q(serial_number__icontains=query) | 
            Q(brand__icontains=query) | 
            Q(model__icontains=query) |
            Q(ip_address__icontains=query)
        )
    
    if area_id:
        equipments_list = equipments_list.filter(area_id=area_id)
    if status:
        equipments_list = equipments_list.filter(status=status)
    if eq_type:
        equipments_list = equipments_list.filter(type=eq_type)
    if ownership:
        equipments_list = equipments_list.filter(ownership_type=ownership)
    
    paginator = Paginator(equipments_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    areas = Area.objects.all()
    from .choices import EQUIPMENT_STATUS_CHOICES, EQUIPMENT_TYPE_CHOICES, OWNERSHIP_CHOICES
        
    context = {
        'page_obj': page_obj, 
        'search_query': query,
        'areas': areas,
        'status_choices': EQUIPMENT_STATUS_CHOICES,
        'type_choices': EQUIPMENT_TYPE_CHOICES,
        'ownership_choices': OWNERSHIP_CHOICES,
        'current_area': int(area_id) if area_id.isdigit() else '',
        'current_status': status,
        'current_type': eq_type,
        'current_ownership': ownership,
    }
    return render(request, 'inventory/equipment_list.html', context)

@login_required
def maintenance_create_view(request):
    if request.method == 'POST':
        form = MaintenanceForm(request.POST)
        if form.is_valid():
            maintenance = form.save(commit=False)
            maintenance.performed_by = request.user
            maintenance.save()
            
            # --- Sync with Schedule ---
            try:
                # Approximate match logic: same equipment, same month/year
                # Or exact match if we want strictly scheduled dates. 
                # Let's check for an existing PENDING schedule in the same month
                start_of_month = maintenance.date.replace(day=1)
                import calendar
                last_day = calendar.monthrange(maintenance.date.year, maintenance.date.month)[1]
                end_of_month = maintenance.date.replace(day=last_day)
                
                schedule = MaintenanceSchedule.objects.filter(
                    equipment=maintenance.equipment,
                    scheduled_date__range=[start_of_month, end_of_month],
                    status='PENDING'
                ).first()
                
                if schedule:
                    # Update existing pending schedule
                    schedule.status = 'COMPLETED'
                    schedule.scheduled_date = maintenance.date # Update date to actual maintenance date
                    schedule.save()
                else:
                    # Create new completed schedule entry if none existed
                    # Check if there is already a completed one for this date to avoid duplicates?
                    exists = MaintenanceSchedule.objects.filter(
                        equipment=maintenance.equipment,
                        scheduled_date=maintenance.date
                    ).exists()
                    
                    if not exists:
                        MaintenanceSchedule.objects.create(
                            equipment=maintenance.equipment,
                            scheduled_date=maintenance.date,
                            status='COMPLETED'
                        )
            except Exception as e:
                print(f"Error syncing schedule: {e}")
            # --------------------------

            return redirect('inventory:maintenance_success', pk=maintenance.pk)
    else:
        form = MaintenanceForm()
            
    return render(request, 'inventory/maintenance_form.html', {'form': form})

@login_required
def equipment_create_view(request):
    if request.method == 'POST':
        form = EquipmentForm(request.POST) 
        if form.is_valid():
            form.save()
            return redirect('inventory:equipment_list')
    else:
        form = EquipmentForm()
    
    return render(request, 'inventory/equipment_form.html', {'form': form, 'title': 'Nuevo Equipo'})

@login_required
def equipment_edit_view(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            form.save()
            return redirect('inventory:equipment_detail', pk=pk)
    else:
        form = EquipmentForm(instance=equipment)
    
    return render(request, 'inventory/equipment_form.html', {'form': form, 'title': f'Editar {equipment.serial_number}'})

@login_required
def area_create_view(request):
    if request.method == 'POST':
        form = AreaForm(request.POST)
        if form.is_valid():
            form.save()
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('inventory:dashboard')
    else:
        form = AreaForm()
    
    return render(request, 'inventory/area_form.html', {'form': form})

@login_required
def area_list_view(request):
    areas = Area.objects.all().order_by('name')
    return render(request, 'inventory/area_list.html', {'areas': areas})

@login_required
def area_edit_view(request, pk):
    area = get_object_or_404(Area, pk=pk)
    if request.method == 'POST':
        form = AreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            return redirect('inventory:area_list')
    else:
        form = AreaForm(instance=area)
    
    return render(request, 'inventory/area_form.html', {'form': form, 'title': f'Editar Área: {area.name}'})

@login_required
def cost_center_create_view(request):
    if request.method == 'POST':
        form = CostCenterForm(request.POST)
        if form.is_valid():
            form.save()
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('inventory:dashboard')
    else:
        form = CostCenterForm()
    
    return render(request, 'inventory/cost_center_form.html', {'form': form})

@login_required
def cost_center_list_view(request):
    cost_centers = CostCenter.objects.all().order_by('code')
    return render(request, 'inventory/cost_center_list.html', {'cost_centers': cost_centers})

@login_required
def cost_center_edit_view(request, pk):
    cost_center = get_object_or_404(CostCenter, pk=pk)
    if request.method == 'POST':
        form = CostCenterForm(request.POST, instance=cost_center)
        if form.is_valid():
            form.save()
            return redirect('inventory:cost_center_list')
    else:
        form = CostCenterForm(instance=cost_center)
    
    return render(request, 'inventory/cost_center_form.html', {'form': form, 'title': f'Editar C.C.: {cost_center.code}'})

@login_required
@user_passes_test(lambda u: u.is_staff)
def user_list_view(request):
    users = User.objects.all().order_by('username')
    return render(request, 'inventory/user_list.html', {'users': users})

@login_required
@user_passes_test(lambda u: u.is_staff)
def user_create_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventory:user_list')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'inventory/user_form.html', {'form': form})

@login_required
def equipment_detail_view(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    maintenances = Maintenance.objects.filter(equipment=equipment).order_by('-date')
    handovers = Handover.objects.filter(equipment=equipment).order_by('-date')
    
    context = {
        'equipment': equipment,
        'maintenances': maintenances,
        'handovers': handovers,
    }
    return render(request, 'inventory/equipment_detail.html', context)

@login_required
def peripheral_list_view(request):
    query = request.GET.get('q', '')
    peripherals_list = Peripheral.objects.all().order_by('-id')
    
    if query:
        peripherals_list = peripherals_list.filter(
            Q(serial_number__icontains=query) | 
            Q(brand__icontains=query) | 
            Q(model__icontains=query) |
            Q(type__icontains=query)
        )

    paginator = Paginator(peripherals_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'inventory/peripheral_list.html', {'page_obj': page_obj, 'search_query': query})

@login_required
def peripheral_detail_view(request, pk):
    peripheral = get_object_or_404(Peripheral, pk=pk)
    handovers = Handover.objects.filter(peripherals=peripheral).order_by('-date')
    
    context = {
        'peripheral': peripheral,
        'handovers': handovers,
    }
    return render(request, 'inventory/peripheral_detail.html', context)

@login_required
def peripheral_create_view(request):
    if request.method == 'POST':
        form = PeripheralForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventory:peripheral_list')
    else:
        form = PeripheralForm()
    
    return render(request, 'inventory/peripheral_form.html', {'form': form, 'title': 'Nuevo Periférico'})

@login_required
def peripheral_edit_view(request, pk):
    peripheral = get_object_or_404(Peripheral, pk=pk)
    if request.method == 'POST':
        form = PeripheralForm(request.POST, instance=peripheral)
        if form.is_valid():
            form.save()
            return redirect('inventory:peripheral_detail', pk=pk)
    else:
        form = PeripheralForm(instance=peripheral)
    
    return render(request, 'inventory/peripheral_form.html', {'form': form, 'title': f'Editar {peripheral.brand} {peripheral.model}'})

@login_required
def client_create_view(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('inventory:dashboard')
    else:
        form = ClientForm()
    
    return render(request, 'inventory/client_form.html', {'form': form})

@login_required
def handover_create_view(request):
    if request.method == 'POST':
        form = HandoverForm(request.POST)
        if form.is_valid():
            # Check if it's a preview or save
            action = request.POST.get('action', 'save')
            
            if action == 'preview':
                # Create a temporary instance without saving
                handover = form.save(commit=False)
                handover.technician = request.user
                handover.date = timezone.now() # Manually set date for preview since auto_now_add won't trigger yet
                
                # Get selected M2M objects directly from cleaned_data
                selected_equipment = form.cleaned_data.get('equipment')
                selected_peripherals = form.cleaned_data.get('peripherals')
                
                # Generate PDF with these explicitly provided lists
                pdf_content = generate_handover_pdf(handover, equipment_list=selected_equipment, peripheral_list=selected_peripherals)
                
                response = HttpResponse(pdf_content, content_type='application/pdf')
                # inline = open in browser, attachment = download
                response['Content-Disposition'] = 'inline; filename="vista_previa_acta.pdf"'
                return response
            
            else:
                # Save action
                handover = form.save(commit=False)
                handover.technician = request.user
                handover.save()
                form.save_m2m() 
                return redirect('inventory:handover_success', pk=handover.pk)
    else:
        form = HandoverForm()
    
    return render(request, 'inventory/handover_form.html', {'form': form, 'title': 'Nueva Entrega / Acta'})

@login_required
def handover_success_view(request, pk):
    return render(request, 'inventory/handover_success.html', {'pk': pk})

@login_required
def maintenance_success_view(request, pk):
    return render(request, 'inventory/maintenance_success.html', {'pk': pk})

@login_required
def maintenance_list_view(request):
    date_start = request.GET.get('date_start', '')
    date_end = request.GET.get('date_end', '')
    m_type = request.GET.get('type', '')

    maintenances = Maintenance.objects.all().order_by('-date')
    
    if date_start:
        maintenances = maintenances.filter(date__gte=date_start)
    if date_end:
        maintenances = maintenances.filter(date__lte=date_end)
    if m_type:
        maintenances = maintenances.filter(maintenance_type=m_type)

    from .choices import MAINTENANCE_TYPE_CHOICES
    
    context = {
        'maintenances': maintenances,
        'maintenance_types': MAINTENANCE_TYPE_CHOICES,
        'current_start': date_start,
        'current_end': date_end,
        'current_type': m_type,
    }
    return render(request, 'inventory/maintenance_list.html', context)

@login_required
def handover_list_view(request):
    date_start = request.GET.get('date_start', '')
    date_end = request.GET.get('date_end', '')
    area_id = request.GET.get('area', '')

    handovers = Handover.objects.all().order_by('-date')
    
    if date_start:
        handovers = handovers.filter(date__date__gte=date_start)
    if date_end:
        handovers = handovers.filter(date__date__lte=date_end)
    if area_id:
        handovers = handovers.filter(Q(source_area_id=area_id) | Q(destination_area_id=area_id))

    areas = Area.objects.all()
    
    context = {
        'handovers': handovers,
        'areas': areas,
        'current_start': date_start,
        'current_end': date_end,
        'current_area': int(area_id) if area_id.isdigit() else '',
    }
    return render(request, 'inventory/handover_list.html', context)


@login_required
def maintenance_schedule_view(request):
    year = int(request.GET.get('year', timezone.now().year))
    
    # Grid structure: [Month 1..12][Week 1..4]
    weeks = [1, 2, 3, 4]
    months = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    equipments = Equipment.objects.all().select_related('area').order_by('area__name', 'serial_number')
    
    # Filter by Area
    area_id = request.GET.get('area')
    if area_id:
        equipments = equipments.filter(area_id=area_id)

    # Fetch schedules for the year
    schedules = MaintenanceSchedule.objects.filter(scheduled_date__year=year)
    
    # Map: (equipment_id, month, visual_week) -> { status, day }
    schedule_map = {}
    for s in schedules:
        day = s.scheduled_date.day
        month = s.scheduled_date.month
        
        # Determine visual week
        if day <= 7: v_week = 1
        elif day <= 14: v_week = 2
        elif day <= 21: v_week = 3
        else: v_week = 4
        
        # Store data. If multiple in same visual week (rare but possible), just overwrite or show one.
        schedule_map[(s.equipment_id, month, v_week)] = {
            'status': s.status,
            'day': day,
            'date': s.scheduled_date.strftime('%Y-%m-%d')
        }
    
    for eq in equipments:
        eq.schedule_row = []
        for m_num, m_name in months:
            month_weeks = []
            for w in weeks:
                data = schedule_map.get((eq.id, m_num, w), None)
                month_weeks.append({'month': m_num, 'week': w, 'data': data})
            eq.schedule_row.append({'month': m_name, 'weeks': month_weeks})
            
    # Generate year range: Current year to +20 years
    current_actual_years = timezone.now().year
    # If the user selected a past year, we might want to include it, but request is "starting 2026 + 20"
    # Let's just do a fixed range relative to NOW or ensure selected year is in it.
    
    start_year = current_actual_years
    available_years = range(start_year, start_year + 21)

    areas = Area.objects.all()

    context = {
        'year': year,
        'months': months,
        'equipments': equipments,
        'weeks': weeks,
        'available_years': available_years,
        'areas': areas,
        'current_area': int(area_id) if area_id and area_id.isdigit() else None,
    }
    return render(request, 'inventory/maintenance_schedule.html', context)

@csrf_exempt
@login_required
def toggle_schedule_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            equipment_id = data.get('equipment_id')
            date_str = data.get('date') # YYYY-MM-DD
            
            if not date_str:
                 return JsonResponse({'error': 'Date required'}, status=400)

            # Check if exists
            schedule = MaintenanceSchedule.objects.filter(
                equipment_id=equipment_id, scheduled_date=date_str
            ).first()
            
            if schedule:
                # Toggle OFF (Delete)
                schedule.delete()
                return JsonResponse({'status': 'removed'})
            else:
                # Toggle ON (Create)
                MaintenanceSchedule.objects.create(
                    equipment_id=equipment_id, scheduled_date=date_str, status='PENDING'
                )
                return JsonResponse({'status': 'added', 'state': 'PENDING'})
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
def reports_dashboard_view(request):
    import datetime
    from django.db.models import Count, Sum
    
    # 1. Date Filtering
    today = timezone.now().date()
    start_str = request.GET.get('start_date', '')
    end_str = request.GET.get('end_date', '')
    
    # Defaults: Start of current year to today
    if not start_str:
        start_date = today.replace(month=1, day=1)
    else:
        try:
            start_date = datetime.datetime.strptime(start_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today.replace(month=1, day=1)
            
    if not end_str:
        end_date = today
    else:
        try:
            end_date = datetime.datetime.strptime(end_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = today

    # 2. Main QuerySets
    maintenances = Maintenance.objects.filter(date__range=[start_date, end_date])
    handovers = Handover.objects.filter(date__date__range=[start_date, end_date])
    equipments = Equipment.objects.all()

    # 3. KPIs
    total_equipment = equipments.count()
    active_equipment = equipments.filter(status='ACTIVE').count()
    maintenance_count = maintenances.count()
    handover_count = handovers.count()
    
    # 4. Charts Data
    
    # 4.1 Equipment by Type
    eq_by_type = list(equipments.values('type').annotate(count=Count('type')).order_by('-count'))
    # Clean up labels from choices? Ideally yes, but values() gives the DB value. 
    # Validating choice labels in template or via dict mapping if needed. For now DB values are readable enough or we format in template.
    
    # 4.2 Maintenance by Type
    m_by_type = list(maintenances.values('maintenance_type').annotate(count=Count('id')))
    
    # 4.3 Top Technicians (by maintenance count)
    top_techs = list(maintenances.values('performed_by__username').annotate(count=Count('id')).order_by('-count')[:5])
    
    # 4.4 Equipment Status
    eq_by_status = list(equipments.values('status').annotate(count=Count('status')))

    # 5. Actionable Lists
    # Warranty Expiring in next 90 days
    limit_date = today + datetime.timedelta(days=90)
    warranty_expiring = equipments.filter(warranty_expiry__range=[today, limit_date]).order_by('warranty_expiry')

    context = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_equipment': total_equipment,
        'active_equipment': active_equipment,
        'maintenance_count': maintenance_count,
        'handover_count': handover_count,
        'eq_by_type': eq_by_type,
        'm_by_type': m_by_type,
        'top_techs': top_techs,
        'eq_by_status': eq_by_status,
        'warranty_expiring': warranty_expiring,
    }
    return render(request, 'inventory/reports_dashboard.html', context)

def export_report_pdf(request):
    import datetime
    from fpdf import FPDF
    from django.http import HttpResponse
    from django.db.models import Count
    
    # 1. Date Filtering (Same logic as dashboard)
    today = timezone.now().date()
    start_str = request.GET.get('start_date', '')
    end_str = request.GET.get('end_date', '')
    
    if not start_str:
        start_date = today.replace(month=1, day=1)
    else:
        try:
            start_date = datetime.datetime.strptime(start_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today.replace(month=1, day=1)
            
    if not end_str:
        end_date = today
    else:
        try:
            end_date = datetime.datetime.strptime(end_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = today

    # 2. Fetch Data
    maintenances = Maintenance.objects.filter(date__range=[start_date, end_date])
    handovers = Handover.objects.filter(date__date__range=[start_date, end_date])
    equipments = Equipment.objects.all()
    
    total_equipment = equipments.count()
    active_equipment = equipments.filter(status='ACTIVE').count()
    maintenance_count = maintenances.count()
    handover_count = handovers.count()
    
    # Warranties expiring next 90 days
    limit_date = today + datetime.timedelta(days=90)
    warranty_expiring = equipments.filter(warranty_expiry__range=[today, limit_date]).order_by('warranty_expiry')

    # 3. Generate PDF
    class PDF(FPDF):
        def header(self):
            self.set_font('Arial', 'B', 15)
            self.cell(0, 10, 'Reporte de Gestión de Inventario y Mantenimiento', 0, 1, 'C')
            self.set_font('Arial', '', 10)
            self.cell(0, 5, f'Generado el: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'C')
            self.ln(10)

        def footer(self):
            self.set_y(-15)
            self.set_font('Arial', 'I', 8)
            self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', 0, 0, 'C')

    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # --- Period Info ---
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f'Periodo del Reporte: {start_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")}', 0, 1, 'L')
    pdf.ln(5)
    
    # --- Executive Summary ---
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, '1. Resumen Ejecutivo', 0, 1, 'L', fill=True)
    pdf.ln(2)
    
    pdf.set_font('Arial', '', 11)
    
    # KPI Grid simulation
    col_width = 45
    pdf.cell(col_width, 10, 'Total Equipos:', 0, 0)
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(col_width, 10, str(total_equipment), 0, 0)
    
    pdf.set_font('Arial', '', 11)
    pdf.cell(col_width, 10, 'Mantenimientos:', 0, 0)
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(col_width, 10, str(maintenance_count), 0, 1)
    
    pdf.set_font('Arial', '', 11)
    pdf.cell(col_width, 10, 'Equipos Activos:', 0, 0)
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(col_width, 10, str(active_equipment), 0, 0)
    
    pdf.set_font('Arial', '', 11)
    pdf.cell(col_width, 10, 'Entregas:', 0, 0)
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(col_width, 10, str(handover_count), 0, 1)
    pdf.ln(5)

    # --- Maintenance Summary ---
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, '2. Resumen de Mantenimientos por Tipo', 0, 1, 'L', fill=True)
    pdf.ln(2)
    
    m_by_type = list(maintenances.values('maintenance_type').annotate(count=Count('id')))
    
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(80, 8, 'Tipo de Mantenimiento', 1)
    pdf.cell(40, 8, 'Cantidad', 1)
    pdf.ln()
    
    # Import choices locally
    from .choices import MAINTENANCE_TYPE_CHOICES
    
    pdf.set_font('Arial', '', 10)
    if m_by_type:
        for item in m_by_type:
            label = dict(MAINTENANCE_TYPE_CHOICES).get(item['maintenance_type'], item['maintenance_type'])
            pdf.cell(80, 8, str(label), 1)
            pdf.cell(40, 8, str(item['count']), 1)
            pdf.ln()
    else:
        pdf.cell(120, 8, 'No se registraron mantenimientos en este periodo.', 1)
        pdf.ln()
    pdf.ln(5)
    
    # --- Top Technicians ---
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, '3. Actividad por Técnico (Top 5)', 0, 1, 'L', fill=True)
    pdf.ln(2)
    
    top_techs = list(maintenances.values('performed_by__username').annotate(count=Count('id')).order_by('-count')[:5])
    
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(80, 8, 'Técnico', 1)
    pdf.cell(40, 8, 'Mantenimientos', 1)
    pdf.ln()
    
    pdf.set_font('Arial', '', 10)
    if top_techs:
        for tech in top_techs:
            username = tech['performed_by__username']
            # Try to get full name if possible, assuming username is available
            user_display = username 
            # (Fetching actual User object for full name would be cleaner but this fits in single view logic)
            pdf.cell(80, 8, str(user_display).title(), 1)
            pdf.cell(40, 8, str(tech['count']), 1)
            pdf.ln()
    else:
        pdf.cell(120, 8, 'Sin actividad registrada.', 1)
        pdf.ln()
    pdf.ln(5)

    # --- Warranty Alerts ---
    if warranty_expiring.exists():
        pdf.add_page()
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(220, 50, 50)
        pdf.cell(0, 10, '4. Alertas de Garantía (Próximos 90 días)', 0, 1, 'L', fill=False)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)
        
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(40, 8, 'Fecha Vence', 1)
        pdf.cell(50, 8, 'Serial', 1)
        pdf.cell(60, 8, 'Equipo', 1)
        pdf.cell(40, 8, 'Área', 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 9)
        for w in warranty_expiring:
            pdf.cell(40, 8, w.warranty_expiry.strftime('%Y-%m-%d'), 1)
            pdf.cell(50, 8, str(w.serial_number), 1)
            pdf.cell(60, 8, f"{w.brand} {w.model}"[:30], 1)
            area_name = w.area.name if w.area else "-"
            pdf.cell(40, 8, area_name[:20], 1)
            pdf.ln()
            
    # --- Detailed Maintenance Log ---
    if maintenances.exists():
        pdf.add_page()
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, '5. Detalle de Mantenimientos Realizados', 0, 1, 'L', fill=True)
        pdf.ln(2)
        
        # Table Header
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(25, 8, 'Fecha', 1)
        pdf.cell(30, 8, 'Serial', 1)
        pdf.cell(35, 8, 'Equipo', 1)
        pdf.cell(30, 8, 'Tipo', 1)
        pdf.cell(35, 8, 'Técnico', 1)
        pdf.cell(35, 8, 'Descripción', 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 8)
        for m in maintenances.select_related('equipment', 'performed_by'):
            pdf.cell(25, 8, m.date.strftime('%Y-%m-%d'), 1)
            pdf.cell(30, 8, m.equipment.serial_number[:15], 1)
            pdf.cell(35, 8, f"{m.equipment.brand} {m.equipment.model}"[:20], 1)
            
            m_type_label = dict(MAINTENANCE_TYPE_CHOICES).get(m.maintenance_type, m.maintenance_type)
            pdf.cell(30, 8, str(m_type_label)[:15], 1)
            
            tech_name = m.performed_by.username if m.performed_by else "N/A"
            pdf.cell(35, 8, tech_name[:18], 1)
            
            # Shorten description
            desc = m.description.replace('\n', ' ')[:20] + "..." if len(m.description) > 20 else m.description
            pdf.cell(35, 8, desc, 1)
            pdf.ln()
            
    # --- Detailed Handover Log ---
    if handovers.exists():
        pdf.add_page()
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, '6. Detalle de Entregas (Actas)', 0, 1, 'L', fill=True)
        pdf.ln(2)
        
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(25, 8, 'Fecha', 1)
        pdf.cell(60, 8, 'Equipos (Serials)', 1) # Widened col
        pdf.cell(35, 8, 'Origen', 1)
        pdf.cell(35, 8, 'Destino', 1)
        pdf.cell(30, 8, 'Tipo', 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 7) # Smaller font for list
        # Use prefetch_related for M2M
        for h in handovers.select_related('source_area', 'destination_area').prefetch_related('equipment'):
            h_date = h.date.strftime('%Y-%m-%d') if hasattr(h.date, 'strftime') else str(h.date)[:10]
            
            # Get list of equipments
            eq_list = [f"{e.serial_number} ({e.brand})" for e in h.equipment.all()]
            eq_str = ", ".join(eq_list)
            
            # MultiCell for equipments potentially
            # Simple approach: truncate if too long or use multicell. Using basic cell for now, truncated.
            
            pdf.cell(25, 8, h_date, 1)
            pdf.cell(60, 8, eq_str[:45], 1)
            
            source = h.source_area.name if h.source_area else "N/A"
            dest = h.destination_area.name if h.destination_area else "N/A"
            
            pdf.cell(35, 8, source[:20], 1)
            pdf.cell(35, 8, dest[:20], 1)
            pdf.cell(30, 8, h.get_type_display()[:15], 1)
            pdf.ln()

    # Output
    response = HttpResponse(content_type='application/pdf')
    # Change attachment to inline for preview
    response['Content-Disposition'] = f'inline; filename="reporte_inventario_{start_date}_{end_date}.pdf"'
    
    # FPDF 1.7.2 in Python 3 returns a string for dest='S'. 
    # Must encode to latin-1 to get the correct binary bytes for the response.
    # Otherwise Django/Python utf-8 encoding corrupts the PDF structure, resulting in a blank or invalid file.
    pdf_content = pdf.output(dest='S').encode('latin-1')
    response.write(pdf_content)
    
    return response
