from datetime import time
import openpyxl
from django.http import HttpResponse
from fpdf import FPDF
import io
import re

def clean_text(text):
    if text is None:
        return ""
    try:
        # Encode to cp1252 (standard Windows encoding often used in PDFs)
        # and decode as latin-1 to keep single-byte characters in Python string
        return str(text).encode('cp1252', 'replace').decode('latin-1')
    except Exception:
        return str(text)


def export_to_excel(queryset, model_admin, request):
    meta = model_admin.model._meta
    field_names = [field.name for field in meta.fields]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\\/*?:\[\]]', '', str(meta.verbose_name_plural))[:31]

    # Header
    ws.append(field_names)

    # Data
    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if value and hasattr(value, 'isoformat'): # Handle dates
                value = value.isoformat()
            if isinstance(value, time):
                value = value.strftime('%H:%M')
            row.append(str(value) if value is not None else '')
        ws.append(row)

    wb.save(response)
    return response

class PDF(FPDF):
    pass

def draw_header(pdf, title, doc_code, date_obj):
    # Generic Header
    pdf.set_font("Arial", size=8)
    start_y = pdf.get_y()
    
    # Logo (Left) - using the SystemSettings or fallback to hfps.jpg
    import os
    from django.conf import settings
    from .models import SystemSettings
    
    sys_settings = SystemSettings.load()
    if sys_settings and sys_settings.logo and os.path.exists(sys_settings.logo.path):
        logo_path = sys_settings.logo.path
    else:
        logo_path = os.path.join(settings.BASE_DIR, 'inventory', 'static', 'img', 'hfps.jpg')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'hfps.jpg')

    if os.path.exists(logo_path):
        pdf.image(logo_path, x=10, y=start_y, w=30, h=20)
        # Assuming A4 width is 210mm, right side image at x=170 + 30w = 200 (10mm margin)
        pdf.image(logo_path, x=170, y=start_y, w=30, h=20)
    else:
        # Fallback if image not found
        pdf.cell(30, 20, "HFPS", border=1, align='C') 
    
    # Title (Center)
    current_x = pdf.get_x()
    pdf.set_xy(current_x, start_y)
    pdf.set_font("Arial", 'B', 10)
    pdf.multi_cell(110, 20, clean_text(title), border=1, align='C') # 20 height to match logo
    
    # Code/Date Block (Right)
    current_x += 110
    pdf.set_xy(current_x, start_y)
    pdf.set_font("Arial", size=8)
    
    # Code
    pdf.cell(50, 5, f"Codigo: {clean_text(doc_code)}", border=1, ln=1)
    pdf.set_x(current_x)
    
    # Date parts
    date_str = date_obj.strftime('%Y-%m-%d')
    day = date_obj.day
    month = date_obj.month
    year = date_obj.year
    
    pdf.cell(50, 5, "Fecha de elaboracion:", border=1, ln=1)
    pdf.set_x(current_x)
    pdf.cell(16, 5, "DD", border=1, align='C')
    pdf.cell(17, 5, "MM", border=1, align='C')
    pdf.cell(17, 5, "AAAA", border=1, ln=1, align='C')
    pdf.set_x(current_x)
    pdf.cell(16, 5, str(day), border=1, align='C')
    pdf.cell(17, 5, str(month), border=1, align='C')
    pdf.cell(17, 5, str(year), border=1, ln=1, align='C')
    
    pdf.set_y(start_y + 25) # Move down past header

def draw_section_title(pdf, title):
    pdf.set_fill_color(220, 220, 220)
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(0, 6, clean_text(title), border=1, ln=1, fill=True)
    pdf.set_font("Arial", size=8)

def draw_field(pdf, label, value, width_label=30, width_value=60, ln=0):
    pdf.set_font("Arial", 'B', 8)
    pdf.cell(width_label, 5, clean_text(label), border=1)
    pdf.set_font("Arial", size=8)
    pdf.cell(width_value, 5, clean_text(str(value) if value else ""), border=1, ln=ln)

def generate_maintenance_pdf(m):
    pdf = PDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    
    draw_header(pdf, "MANTENIMIENTO PREVENTIVO Y CORRECTIVO \nEQUIPOS DE COMPUTO", "FR-GT-03 V.01", m.date)

    # ... 1. DATOS DEL EQUIPO ...
    draw_section_title(pdf, "1. DATOS DEL EQUIPO")
    draw_field(pdf, "Propiedad:", "HFPS", 30, 30)
    draw_field(pdf, "Tipo adquisicion:", "Propio", 30, 30)
    draw_field(pdf, "Estado:", clean_text(m.equipment.get_status_display()), 20, 50, ln=1)
    
    draw_field(pdf, "No. Serie (S/N):", clean_text(m.equipment.serial_number), 30, 60)
    draw_field(pdf, "Nombre Equipo:", clean_text(f"{m.equipment.brand} {m.equipment.model}"), 30, 70, ln=1)
    
    draw_field(pdf, "NomUsuarioSO:", clean_text(m.equipment.os_user or ""), 30, 60)
    pdf.cell(100, 5, "", border=1, ln=1) # Spacer

    pdf.ln(2)

    # ... 2. CENTRO COSTOS ...
    draw_section_title(pdf, "2. INFORMACION CENTRO DE COSTOS")
    area = m.equipment.area
    cc = area.cost_center if area and area.cost_center else None
    
    draw_field(pdf, "Centro De Costos:", clean_text(f"{cc.code} - {cc.name}" if cc else "N/A"), 40, 150, ln=1)
    draw_field(pdf, "Ubicacion Default:", clean_text(area.name if area else "N/A"), 40, 150, ln=1)
    draw_field(pdf, "Proceso/Area:", clean_text(area.description or area.name if area else ""), 40, 150, ln=1)
    
    pdf.ln(2)
    
    # ... 3. CONFIGURACION ...
    draw_section_title(pdf, "3. CONFIGURACION DEL EQUIPO")
    draw_field(pdf, "Tipo de Equipo:", clean_text(m.equipment.get_type_display()), 30, 40)
    draw_field(pdf, "Marca:", clean_text(m.equipment.brand), 20, 40)
    draw_field(pdf, "Modelo:", clean_text(m.equipment.model), 20, 40, ln=1)
    
    draw_field(pdf, "Voltaje:", clean_text(m.equipment.voltage), 30, 40)
    draw_field(pdf, "Amperaje:", clean_text(m.equipment.amperage), 20, 40)
    draw_field(pdf, "Sist. Operativo:", clean_text(m.equipment.operating_system), 20, 40, ln=1)
    
    draw_field(pdf, "Tamano Pantalla:", clean_text(m.equipment.screen_size), 40, 150, ln=1)
    
    pdf.ln(2)

    # ... 4. SUPPORT ...
    draw_section_title(pdf, "4. TIPO DE SOPORTE REQUERIDO")
    # ... logic for checkboxes ...
    # Keep existing logic for checkboxes but clean up if needed.
    pdf.set_font("Arial", size=8)
    pdf.cell(47, 5, f"[ {'X' if m.type_review else ' '} ] REVISION DE EQUIPO", border=1)
    pdf.cell(47, 5, f"[ {'X' if m.type_software_failure else ' '} ] FALLAS CON SOFTWARE", border=1)
    pdf.cell(47, 5, f"[ {'X' if m.type_connection else ' '} ] PROBLEMAS CONEXION", border=1)
    pdf.cell(49, 5, f"[ {'X' if m.type_updates else ' '} ] ACTUALIZACIONES WIN10", border=1, ln=1)

    pdf.cell(47, 5, f"[ {'X' if m.type_cleaning else ' '} ] LIMPIEZA EQUIPO", border=1)
    pdf.cell(47, 5, f"[ {'X' if m.type_install else ' '} ] INSTALACION SOFTWARE", border=1)
    pdf.cell(47, 5, f"[ {'X' if m.type_peripheral else ' '} ] FALLA PERIFERICOS", border=1)
    pdf.cell(49, 5, f"[ {'X' if m.type_backup else ' '} ] COPIA SEGURIDAD", border=1, ln=1)

    pdf.ln(2)
    
    # ... 5. CLEANING ...
    draw_section_title(pdf, "5. ACTIVIDADES DE LIMPIEZA AL SISTEMA OPERATIVO")
    pdf.cell(95, 5, f"[ {'X' if m.cleaning_defrag else ' '} ] DESFRAGMENTACION DISCO DURO", border=1)
    pdf.cell(95, 5, f"[ {'X' if m.cleaning_cco else ' '} ] EJECUCION C-CLEANER", border=1, ln=1)
    pdf.cell(95, 5, f"[ {'X' if m.cleaning_scandisk else ' '} ] SCANDISK", border=1)
    pdf.cell(95, 5, f"[ {'X' if m.cleaning_space else ' '} ] LIBERACION DE ESPACIO", border=1, ln=1)
    
    pdf.ln(2)

    # ... 6. OBSERVATIONS ...
    draw_section_title(pdf, "6. OBSERVACIONES DEL EQUIPO ANTES DEL MANTENIMIENTO")
    pdf.multi_cell(0, 5, clean_text(m.description or "N/A"), border=1)
    
    pdf.ln(2)
    
    # ... 7. HARDWARE ...
    draw_section_title(pdf, "7. ETAPAS DEL MANTENIMIENTO DEL HARDWARE")
    hw_tasks = [
        ("Retiro del equipo / Desmontaje", m.hw_disassembly),
        ("Limpieza Fuente de Poder", m.hw_power_supply),
        ("Limpieza Ventiladores", m.hw_fans),
        ("Limpieza Chasis y Cubiertas", m.hw_chassis),
        ("Uso Crema Disipadora", m.hw_thermal_paste),
        ("Uso Limpiador Contactos", m.hw_contacts),
        ("Limpieza Teclado y Mouse", m.hw_keyboard_mouse),
        ("Limpieza Monitor/Pantalla", m.hw_screen),
        ("Ensamble y Pruebas", m.hw_reassembly)
    ]
    for task, checked in hw_tasks:
        pdf.cell(95, 5, task, border=1)
        pdf.cell(20, 5, "REALIZADO" if checked else "N/A", border=1, align='C', ln=1)

    pdf.ln(5)
    
    # FOOTER
    draw_section_title(pdf, "ESTADO FINAL Y TIEMPOS")
    pdf.cell(40, 5, "ESTADO FINAL:", border=1)
    pdf.cell(150, 5, clean_text(m.equipment.get_status_display()), border=1, ln=1)
    
    start_time = m.start_time.strftime('%H:%M') if m.start_time else "  :  "
    end_time = m.end_time.strftime('%H:%M') if m.end_time else "  :  "
    pdf.cell(40, 5, "HORA INICIO:", border=1)
    pdf.cell(55, 5, start_time, border=1)
    pdf.cell(40, 5, "HORA FINAL:", border=1)
    pdf.cell(55, 5, end_time, border=1, ln=1)

    pdf.ln(10)

    # Signatures
    y_sig = pdf.get_y()
    pdf.line(10, y_sig + 15, 90, y_sig + 15)
    pdf.set_xy(10, y_sig + 16)
    pdf.cell(80, 5, "NOMBRE USUARIO / RECIBIDO POR", align='C', ln=1)
    
    pdf.line(110, y_sig + 15, 190, y_sig + 15)
    pdf.set_xy(110, y_sig + 16)
    tech_name = m.performed_by.get_full_name() if m.performed_by else "TECNICO"
    pdf.cell(80, 5, clean_text(tech_name), align='C', ln=1)
    pdf.set_x(110)
    pdf.cell(80, 5, "SOPORTE TECNICO - TIC", align='C', ln=1)

    # Output
    pdf_content = pdf.output(dest='S')
    if isinstance(pdf_content, str):
        return pdf_content.encode('latin-1')
    return bytes(pdf_content)


def generate_handover_pdf(handover, equipment_list=None, peripheral_list=None):
    pdf = PDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    
    draw_header(pdf, "ACTA DE ENTREGA / ASIGNACION\nEQUIPOS DE COMPUTO Y PERIFERICOS", "FR-GT-04 V.01", handover.date)

    # ... 1. DATOS DE LA ENTREGA ...
    draw_section_title(pdf, "1. INFORMACION DE LA ENTREGA")
    
    draw_field(pdf, "Tipo de Entrega:", clean_text(handover.get_type_display()), 40, 150, ln=1)
    draw_field(pdf, "Area Origen:", clean_text(str(handover.source_area)), 40, 55)
    draw_field(pdf, "Area Destino:", clean_text(str(handover.destination_area)), 40, 55, ln=1)
    
    # Client Info
    if handover.client:
        draw_field(pdf, "Funcionario/Cliente:", clean_text(handover.client.name), 40, 150, ln=1)
        draw_field(pdf, "Identificacion:", clean_text(handover.client.identification), 40, 55)
        draw_field(pdf, "Cargo/Correo:", clean_text(handover.client.email or ""), 40, 55, ln=1)
    else:
        draw_field(pdf, "Recibido Por (Nombre):", clean_text(handover.receiver_name or "N/A"), 40, 150, ln=1)

    pdf.ln(2)

    # ... 2. EQUIPOS ...
    draw_section_title(pdf, "2. EQUIPOS RELACIONADOS")
    
    # Checkbox header? Or just a table
    pdf.set_font("Arial", 'B', 8)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(30, 6, "TIPO", border=1, fill=True)
    pdf.cell(40, 6, "MARCA", border=1, fill=True)
    pdf.cell(50, 6, "MODELO", border=1, fill=True)
    pdf.cell(70, 6, "SERIAL / CODIGO", border=1, ln=1, fill=True)
    
    pdf.set_font("Arial", size=8)
    
    final_equipment = equipment_list if equipment_list is not None else handover.equipment.all()
    
    if not final_equipment:
        pdf.cell(190, 6, "No aplica / Ninguno", border=1, ln=1, align='C')
    else:
        for eq in final_equipment:
            pdf.cell(30, 6, clean_text(eq.get_type_display()[:15]), border=1)
            pdf.cell(40, 6, clean_text(eq.brand[:20]), border=1)
            pdf.cell(50, 6, clean_text(eq.model[:25]), border=1)
            pdf.cell(70, 6, clean_text(eq.serial_number), border=1, ln=1)

    pdf.ln(2)

    # ... 3. PERIFERICOS ...
    draw_section_title(pdf, "3. PERIFERICOS / ACCESORIOS")
    
    pdf.set_font("Arial", 'B', 8)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(30, 6, "TIPO", border=1, fill=True)
    pdf.cell(35, 6, "MARCA", border=1, fill=True) # Reduced
    pdf.cell(45, 6, "MODELO", border=1, fill=True) # Reduced
    pdf.cell(15, 6, "CANT", border=1, fill=True)   # Added
    pdf.cell(65, 6, "SERIAL / ESTADO", border=1, ln=1, fill=True) # Reduced
    
    pdf.set_font("Arial", size=8)
    
    # We expect peripheral_list to be a list of HandoverPeripheral-like objects (having .peripheral and .quantity)
    # If not provided, fetch from DB using the through model to get quantity
    if peripheral_list is not None:
        final_peripherals = peripheral_list
    else:
        # Use the through model manager
        # If handover is saved, we can query HandoverPeripheral
        if handover.pk:
            from .models import HandoverPeripheral
            final_peripherals = HandoverPeripheral.objects.filter(handover=handover)
        else:
            final_peripherals = [] # Should not happen unless unsaved handover without list

    if not final_peripherals:
         pdf.cell(190, 6, "No aplica / Ninguno", border=1, ln=1, align='C')
    else:
        for hp in final_peripherals:
            # hp might be a HandoverPeripheral instance OR a dict/object from preview
            # normalize access
            if hasattr(hp, 'peripheral'):
                p = hp.peripheral
                qty = hp.quantity
            else:
                # Fallback if passed raw Peripheral objects (should avoid this)
                p = hp
                qty = 1
            
            serial = p.serial_number if p.serial_number else p.get_status_display()
            pdf.cell(30, 6, clean_text(str(p.type)[:15]), border=1)
            pdf.cell(35, 6, clean_text(p.brand[:20]), border=1)
            pdf.cell(45, 6, clean_text(p.model[:25]), border=1)
            pdf.cell(15, 6, str(qty), border=1, align='C')
            pdf.cell(65, 6, clean_text(serial), border=1, ln=1)

    pdf.ln(2)
    
    # ... 4. OBSERVACIONES ...
    draw_section_title(pdf, "4. OBSERVACIONES")
    pdf.multi_cell(0, 5, clean_text(handover.observations or "Sin observaciones adicionales."), border=1)
    
    pdf.ln(20) # Space for signatures

    # Signatures
    y_sig = pdf.get_y()
    
    # Check if page break needed
    if y_sig > 250:
        pdf.add_page()
        y_sig = pdf.get_y() + 20
    
    # Area de firmas
    pdf.set_font("Arial", 'B', 8)
    
    # Left: Deliverer (Technician usually)
    pdf.line(20, y_sig, 90, y_sig)
    pdf.set_xy(20, y_sig + 1)
    tech_name = handover.technician.get_full_name() if handover.technician else "TECNICO / RESPONSABLE"
    pdf.cell(70, 5, "ENTREGADO POR:", align='C', ln=1)
    pdf.set_x(20)
    pdf.set_font("Arial", '', 8)
    pdf.cell(70, 5, clean_text(tech_name), align='C', ln=1)
    
    # Right: Receiver
    pdf.line(120, y_sig, 190, y_sig)
    pdf.set_xy(120, y_sig + 1)
    pdf.set_font("Arial", 'B', 8)
    pdf.cell(70, 5, "RECIBIDO POR:", align='C', ln=1)
    
    receiver_label = handover.receiver_name
    if handover.client:
        receiver_label = handover.client.name
        
    pdf.set_x(120)
    pdf.set_font("Arial", '', 8)
    pdf.cell(70, 5, clean_text(receiver_label or "_________"), align='C', ln=1)

    # Output
    pdf_content = pdf.output(dest='S')
    if isinstance(pdf_content, str):
        return pdf_content.encode('latin-1')
    return bytes(pdf_content)

def generate_equipment_history_pdf(equipment, events):
    from django.utils import timezone
    pdf = PDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    
    draw_header(pdf, "HOJA DE VIDA Y BITACORA HISTORICA\nEQUIPOS DE COMPUTO", "FR-GT-HV V.01", timezone.now().date())

    # ... 1. DATOS DEL EQUIPO ...
    draw_section_title(pdf, "1. IDENTIFICACION PRINCIPAL")
    
    draw_field(pdf, "Tipo de Equipo:", clean_text(equipment.get_type_display()), 30, 40)
    draw_field(pdf, "Marca:", clean_text(equipment.brand), 20, 40)
    draw_field(pdf, "Modelo:", clean_text(equipment.model), 20, 40, ln=1)
    
    draw_field(pdf, "No. Serie / ID:", clean_text(equipment.serial_number), 30, 40)
    draw_field(pdf, "Estado Actual:", clean_text(equipment.get_status_display()), 20, 40)
    
    area_name = equipment.area.name if equipment.area else "Sin asignar"
    draw_field(pdf, "Ubicacion:", clean_text(area_name), 20, 40, ln=1)

    pdf.ln(2)

    # ... 2. ESPECIFICACIONES TECNICAS ...
    draw_section_title(pdf, "2. ESPECIFICACIONES TECNICAS")
    
    draw_field(pdf, "Sis. Operativo:", clean_text(equipment.operating_system or "-"), 30, 40)
    draw_field(pdf, "Procesador:", clean_text(equipment.processor or "-"), 20, 40)
    draw_field(pdf, "Memoria RAM:", clean_text(equipment.ram or "-"), 20, 40, ln=1)
    
    draw_field(pdf, "Almacenamiento:", clean_text(equipment.storage or "-"), 30, 40)
    draw_field(pdf, "Direccion IP:", clean_text(equipment.ip_address or "-"), 20, 40)
    draw_field(pdf, "MAC Address:", clean_text(equipment.mac_address or "-"), 20, 40, ln=1)

    pdf.ln(2)

    # ... 3. LOGISTICA Y GARANTIA ...
    draw_section_title(pdf, "3. LOGISTICA Y GARANTIA")
    
    ownership_str = str(equipment.ownership) if equipment.ownership else "Propio"
    draw_field(pdf, "Tipo Propiedad:", clean_text(ownership_str), 30, 40)
    draw_field(pdf, "Fecha Compra:", clean_text(str(equipment.purchase_date or "-")), 25, 35)
    draw_field(pdf, "Fin Vida Util:", clean_text(str(equipment.end_of_life_date or "-")), 25, 35, ln=1)

    # ... 4. TIMELINE EVENTS ...
    pdf.ln(5)
    draw_section_title(pdf, "4. HISTORIAL Y MOVIMIENTOS (LINEA DE TIEMPO)")
    
    if not events:
        pdf.set_font("Arial", 'I', 9)
        pdf.cell(0, 8, "No existen registros previos en el sistema para este equipo.", ln=1)
    else:
        pdf.set_font("Arial", 'B', 8)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(35, 6, "FECHA / HORA", border=1, fill=True)
        pdf.cell(35, 6, "TIPO EVENTO", border=1, fill=True)
        pdf.cell(80, 6, "DESCRIPCION / DETALLES", border=1, fill=True)
        pdf.cell(40, 6, "USUARIO", border=1, ln=1, fill=True)
        
        pdf.set_font("Arial", size=8)
        
        for e in events:
            # Check Y position to avoid drawing off-page
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_font("Arial", 'B', 8)
                pdf.set_fill_color(240, 240, 240)
                pdf.cell(35, 6, "FECHA / HORA", border=1, fill=True)
                pdf.cell(35, 6, "TIPO EVENTO", border=1, fill=True)
                pdf.cell(80, 6, "DESCRIPCION / DETALLES", border=1, fill=True)
                pdf.cell(40, 6, "USUARIO", border=1, ln=1, fill=True)
                pdf.set_font("Arial", size=8)

            date_str = e['date'].strftime('%Y-%m-%d %H:%M') if e['date'] else "-"
            tipo_map = {
                'timeline': "SISTEMA",
                'maintenance': "MANTENIMIENTO",
                'handover': "ACTA ENTREGA",
                'round': "RONDA",
                'component': "MOD. COMPONENTE",
                'retirement': "BAJA"
            }
            tipo_label = tipo_map.get(e['type'], e['type'].upper())
            
            # Use MultiCell for description in case it's long, which requires tracking X/Y manually for row
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            
            desc_text = clean_text(e['title'] + " - " + str(e['description']))
            user_text = clean_text(str(e['user']))
            
            # Predict height needed for description
            pdf.set_font("Arial", size=8)
            lines = pdf.get_string_width(desc_text) / 80.0
            h = max(6, int(max(1, lines) * 4) + 2)
            
            # Handle page break inside row logic roughly
            if y_start + h > 275:
                 pdf.add_page()
                 x_start = pdf.get_x()
                 y_start = pdf.get_y()
                 pdf.set_font("Arial", size=8)
            
            pdf.rect(x_start, y_start, 35, h)
            pdf.rect(x_start + 35, y_start, 35, h)
            pdf.rect(x_start + 70, y_start, 80, h)
            pdf.rect(x_start + 150, y_start, 40, h)
            
            pdf.set_xy(x_start, y_start)
            pdf.cell(35, h, date_str, align='C')
            pdf.cell(35, h, tipo_label, align='C')
            
            pdf.set_xy(x_start + 70, y_start + 1)
            pdf.multi_cell(80, 4, desc_text, align='L')
            
            pdf.set_xy(x_start + 150, y_start)
            pdf.cell(40, h, user_text, align='C')
            
            pdf.set_xy(x_start, y_start + h)

    pdf.ln(15)

    # Footer Signatures
    y_sig = pdf.get_y()
    if y_sig > 250:
        pdf.add_page()
        y_sig = pdf.get_y() + 20

    pdf.set_font("Arial", 'B', 8)
    pdf.line(60, y_sig, 150, y_sig)
    pdf.set_xy(60, y_sig + 1)
    pdf.cell(90, 5, "COORDINACION TIC - SISTEMAS", align='C', ln=1)
    
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    pdf.set_font("Arial", 'I', 7)
    pdf.set_xy(60, y_sig + 6)
    pdf.cell(90, 5, f"Generado por Sistema el {timestamp}", align='C', ln=1)

    # Output
    pdf_content = pdf.output(dest='S')
    buffer = io.BytesIO(pdf_content.encode('latin-1') if isinstance(pdf_content, str) else bytes(pdf_content))
    filename = f"hoja-de-vida-{equipment.serial_number or equipment.pk}.pdf"
    
    return buffer, filename
